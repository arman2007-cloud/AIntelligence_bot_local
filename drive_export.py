import os
import sys
import datetime
import tempfile
import re
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()

drive_folder_id = os.getenv("DRIVE_FOLDER_ID", "")
drive_available = False

try:
    from google.oauth2.credentials import Credentials
    from google_auth_oauthlib.flow import InstalledAppFlow
    from google.auth.transport.requests import Request
    from googleapiclient.discovery import build
    from googleapiclient.http import MediaFileUpload
    drive_available = True
except ImportError:
    pass

def get_resource_path(filename):
    if getattr(sys, 'frozen', False):
        base_directory = os.path.dirname(sys.executable)
    else:
        base_directory = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_directory, filename)

scopes = ["https://www.googleapis.com/auth/drive.file"]
token_path = get_resource_path("drive_token.json")
credentials_path = get_resource_path("credentials.json")

def get_drive_service():
    if not drive_available:
        raise ImportError("Google API libraries are missing.")

    google_credentials = None
    if os.path.exists(token_path):
        google_credentials = Credentials.from_authorized_user_file(token_path, scopes)

    if not google_credentials or not google_credentials.valid:
        if google_credentials and google_credentials.expired and google_credentials.refresh_token:
            google_credentials.refresh(Request())
        else:
            if not os.path.exists(credentials_path):
                raise FileNotFoundError(f"Alert: File '{credentials_path}' not found.")
            oauth_flow = InstalledAppFlow.from_client_secrets_file(credentials_path, scopes)
            google_credentials = oauth_flow.run_local_server(port=0)

        with open(token_path, "w") as token_file:
            token_file.write(google_credentials.to_json())

    return build("drive", "v3", credentials=google_credentials)

def build_excel(data, keyword, location, mode="leads"):
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Job Market Radar" if mode == "jobs" else "Talent Pipeline"

    header_color = "FF07070C"
    high_score_color = "FF00C896"
    mid_score_color = "FFFFB347"
    low_score_color = "FFF43F5E"
    row_color_even = "FF141420"
    row_color_odd = "FF0F0F17"
    text_color = "FFEEEEF5"
    link_color = "FF00D4FF"

    if mode == "jobs":
        columns_headers = ["Score", "Company", "Job Title", "Location", "Job URL"]
        column_widths = [9, 30, 50, 30, 60]
    else:
        columns_headers = ["Score", "Name", "Job Title", "Company", "Location", "LinkedIn URL"]
        column_widths = [9, 30, 40, 30, 16, 60]

    cell_border = Border(
        left=Side(border_style="thin", color="FF252535"),
        right=Side(border_style="thin", color="FF252535"),
        top=Side(border_style="thin", color="FF252535"),
        bottom=Side(border_style="thin", color="FF252535"),
    )

    for col_index, (header_text, width) in enumerate(zip(columns_headers, column_widths), 1):
        header_cell = worksheet.cell(row=1, column=col_index, value=header_text)
        header_cell.font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
        header_cell.fill = PatternFill("solid", fgColor=header_color)
        header_cell.alignment = Alignment(horizontal="center", vertical="center")
        header_cell.border = cell_border
        worksheet.column_dimensions[get_column_letter(col_index)].width = width

    for row_index, current_item in enumerate(data, 2):
        item_score = int(current_item.get("score", 0))
        active_score_color = high_score_color if item_score >= 70 else mid_score_color if item_score >= 40 else low_score_color
        active_row_color = row_color_even if row_index % 2 == 0 else row_color_odd

        if mode == "jobs":
            row_values = [item_score, current_item.get("company", "Confidencial"), current_item.get("job_title", "Unknown"), current_item.get("location", location), current_item.get("url", "")]
            url_column_index = 5
        else:
            row_values = [item_score, current_item.get("name", ""), current_item.get("job_title", ""), current_item.get("company", ""), current_item.get("location", location), current_item.get("url", "")]
            url_column_index = 6

        for col_index, cell_value in enumerate(row_values, 1):
            data_cell = worksheet.cell(row=row_index, column=col_index, value=cell_value)
            data_cell.font = Font(color=text_color, name="Segoe UI", size=10)
            data_cell.fill = PatternFill("solid", fgColor=active_row_color)
            data_cell.border = cell_border
            data_cell.alignment = Alignment(vertical="center")

        score_cell = worksheet.cell(row=row_index, column=1)
        score_cell.fill = PatternFill("solid", fgColor=active_score_color)
        score_cell.font = Font(bold=True, color="000000")
        score_cell.alignment = Alignment(horizontal="center", vertical="center")

        url_string = str(current_item.get("url", ""))
        if url_string.startswith("http"):
            link_cell = worksheet.cell(row=row_index, column=url_column_index)
            link_cell.hyperlink = url_string
            link_cell.font = Font(color=link_color, underline="single")

    worksheet.freeze_panes = "A2"

    safe_keyword = re.sub(r'[<>:"/\\|?*]', '', keyword).replace(" ", "_")[:20]
    file_name = f"{'Jobs' if mode == 'jobs' else 'Leads'}_{safe_keyword}_{datetime.date.today()}.xlsx"
    absolute_file_path = os.path.join(tempfile.gettempdir(), file_name)
    workbook.save(absolute_file_path)
    return absolute_file_path

def export_and_upload(keyword, location, min_score=0, leads=None, mode="jobs", log_fn=print):
    if not leads:
        return ""

    generated_excel_path = build_excel(leads, keyword, location, mode)

    if not drive_available or not os.path.exists(credentials_path) or not drive_folder_id:
        log_fn(f"Google Drive is not configured. File saved locally: {generated_excel_path}")
        return generated_excel_path

    try:
        drive_service = get_drive_service()
        clean_file_name = os.path.basename(generated_excel_path).replace('.xlsx', '')

        file_metadata = {
            "name": clean_file_name,
            "parents": [drive_folder_id],
            "mimeType": "application/vnd.google-apps.spreadsheet"
        }

        file_media = MediaFileUpload(
            generated_excel_path,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            resumable=False
        )

        uploaded_file = drive_service.files().create(body=file_metadata, media_body=file_media, fields="webViewLink").execute()
        return uploaded_file.get("webViewLink", "")

    except Exception as error_message:
        log_fn(f"Drive Error: {error_message}")
        return ""
    finally:
        if os.path.exists(generated_excel_path):
            try:
                os.remove(generated_excel_path)
            except:
                pass